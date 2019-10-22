# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import sys

#print(sys.argv[0])
#print(sys.argv[1])
#print(sys.argv[2])

wb = load_workbook("../data/example.xlsx")

print(wb.sheetnames)

ws = wb['Sheet1']

print(ws.cell(row=1, column=3).value)

#for row in range(1,106):
#    for col in range(1, 4):
#    print("%d %d" % (row, col))

for row in ws.iter_rows(min_row=3, max_row=5, min_col=2, max_col=5):
    print(*[r.value for r in row])

